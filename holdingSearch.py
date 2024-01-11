# -*- coding = utf-8 -*-
# @Time  : 2023/3/13 16:09
# @Author: Ifory
# @File  : aqc.py

import argparse
import json
import os
import random
import re
import warnings
from datetime import datetime
from time import sleep

import openpyxl
import requests
from openpyxl.styles import NamedStyle, Font, Alignment

warnings.filterwarnings("ignore")


class EnterInfoSearch(object):
    def __init__(self):
        self.company = None
        self.cookie = ''
        self.pid = None
        self.c_name = None
        self.data_list = []

    def build_headers(self):
        if os.path.exists("cookie.config"):
            with open("cookie.config", "r") as f:
                self.cookie = f.readline()
        else:
            print("cookie.config配置文件不存在！")

        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
            '(KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 '
            '(KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 '
            '(KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:54.0) Gecko/20100101 Firefox/68.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.13; rv:61.0) '
            'Gecko/20100101 Firefox/68.0',
            'Mozilla/5.0 (X11; Linux i586; rv:31.0) Gecko/20100101 Firefox/68.0'
        ]
        ua = random.choice(user_agents)
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Connection': 'Keep-Alive',
            'Cookie': self.cookie,
            'Referer': "https://aiqicha.baidu.com/company_detail_28684192501916",
            'User-Agent': ua
        }
        return headers

    # 统一代理请求
    def get_req(self, url, redirect, is_json=False, t=0):
        if t > 20:
            print("失败请求超过20次 {}".format(url))
            raise Exception(print("请求错误尝试超过20次，自动退出"))
        try:
            resp = requests.get(url, headers=self.build_headers(), verify=False, timeout=10,
                                allow_redirects=redirect,
                                )
            if resp.status_code == 200:
                res = resp.text
                if "百度安全验证" in resp.content.decode():
                    print("百度安全验证，cookie失效或需图形码验证")
                    print("需手动刷新网页验证码或更换cookie")
                    if input("是否刷新完成? y/n  ") != 'y':
                        exit(0)
                else:
                    if is_json:

                        if resp.json()['status'] != 0:
                            print("JSON校验错误重试返回内容:  {} ".format(res))
                            return self.get_req(url, redirect, is_json, t + 1)
                return res
            elif resp.status_code == 302:
                print("风险校验需要更新Cookie {}".format(url))
                return None
            else:
                return self.get_req(url, redirect, is_json, t + 1)
        except requests.exceptions.Timeout:
            print("连接超时自动重连")
            sleep(1)
            return self.get_req(url, redirect, is_json, t + 1)
        except Exception as e:
            print("请求错误 {} ".format(e))
            return None

    def parse_index(self, content, flag=True):
        tag_1 = 'window.pageData ='
        tag_2 = 'window.isSpider ='
        idx_1 = content.find(tag_1)
        idx_2 = content.find(tag_2)
        # 判断企业区间中的JSON数据来进行匹配
        if idx_2 > idx_1:
            # 企业提取判断，去除多余字符
            mystr = content[idx_1 + len(tag_1): idx_2].strip()
            mystr = mystr.replace("\n", "")
            mystr = mystr.replace("window.isSpider = null;", "")
            mystr = mystr.replace("window.updateTime = null;", "")
            mystr = mystr.replace(" ", "")
            mystr = mystr.replace(
                "if(window.pageData.result.isDidiwei){window.location.href=`/login?u=${encodeURIComponent(window.location.href)}`}",
                "")
            mystr = mystr.replace(" ", "")

            len_str = len(mystr)
            if mystr[len_str - 1] == ';':
                mystr = mystr[0:len_str - 1]
            j = json.loads(mystr)
            if flag:
                return j["result"]
            if len(j["result"]["resultList"]) > 0:
                item = j["result"]["resultList"][0]
                return item
            else:
                return None
        else:
            print("企业数据提取失败 {}".format(idx_1))
            return None

    def get_item_name(self, item):
        entName = item['entName']
        pattern = re.compile(r'<[^>]+>', re.S)
        result = pattern.sub('', entName)
        return item['pid'], result

    # 查询企业信息
    def get_cm_if(self, name, t=0):
        """获取PID"""
        company = name
        item = None
        url_prefix = 'https://www.baidu.com/'
        url_a = 'https://aiqicha.baidu.com/s?q=' + company + '&t=0'
        content = self.get_req(url_a, url_prefix, False)
        # print(content)
        if content:
            item = self.parse_index(content, False)
        if t > 3:
            return None
        if item:
            return item
        else:
            print("企业查询重试: {}  {}".format(name, t))
            return self.get_cm_if(name, t + 1)

    def check_name(self):
        item = self.get_cm_if(self.company)
        # print(item)
        if item:
            my = self.get_item_name(item)
            if self.c_name is None:
                self.c_name = my[1]
            return my
        else:
            print("未查询到企业: {}".format(self.company))
            return None

    def getHoldingData(self):
        """获取控股企业数据"""
        print("获取控股企业数据: ")
        res = []
        p = 1
        while True:
            url = f"https://aiqicha.baidu.com:443/detail/holdsAjax?pid={self.pid}&p={p}&size=30&confirm="
            req = self.get_req(url, False)
            # print(req)
            data = json.loads(req)['data']['list']
            if not data:
                print(f"共获取控股企业数据 ({len(res)}个): ")
                break
            res = res + data
            for name in data:
                entName = name['entName']
                proportion = name['proportion']
                if entName and proportion:
                    print("企业名称: {:<50}{:<20}".format(entName, "股权占比: " + str(proportion)))
                    self.data_list.append((entName, proportion, self.company))
            p += 1
        # print(res)

    def save_excel(self, now_time):
        if not os.path.exists("res"):
            os.mkdir("res")
        file_path = f"res/{now_time}.xlsx"
        TitleStyle = NamedStyle(name='TitleStyle',
                                font=Font(name='宋体', size=11, bold=True),
                                alignment=Alignment(horizontal='left', vertical='center'))

        BodyStyle = NamedStyle(name='BodyStyle',
                               alignment=Alignment(horizontal='left', vertical='center'))

        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet("控股企业", 0)

            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 10
            worksheet.column_dimensions['C'].width = 50
            col = ['公司名', '股权占比', '查询企业名称']
            worksheet.append(col)
            for res in self.data_list:
                worksheet.append(res)

            # 标题样式
            for c in range(1, 4):
                worksheet.cell(1, c).style = TitleStyle
            worksheet.freeze_panes = 'A2'  # 冻结首行

            # 正文样式
            for r in range(2, len(self.data_list) + 2):
                worksheet.cell(r, 3).style = BodyStyle

            workbook.save(filename=file_path)
        else:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            max_row = worksheet.max_row  # 获得行数
            for res in self.data_list:
                worksheet.append(res)
            for r in range(max_row + 1, max_row + len(self.data_list) + 1):
                worksheet.cell(r, 3)

            workbook.save(file_path)

    def run(self, now_time, company):
        self.company = company
        print("\n查询企业名称: {}".format(self.company))
        res = self.check_name()
        if res is not None:
            self.pid = res[0]
        if self.pid is not None:
            print("企业PID:      {}".format(self.pid))
            self.getHoldingData()  # 查询控股企业
            self.save_excel(now_time)
            self.data_list = []

    def banner(self):
        print("""
 _           _     _ _              ______                        _     
| |         | |   | (_)            / _____)                      | |    
| |__   ___ | | __| |_ ____   ____( (____  _____ _____  ____ ____| |__  
|  _ \ / _ \| |/ _  | |  _ \ / _  |\____ \| ___ (____ |/ ___) ___)  _ \ 
| | | | |_| | ( (_| | | | | ( (_| |_____) ) ____/ ___ | |  ( (___| | | |
|_| |_|\___/ \_)____|_|_| |_|\___ (______/|_____)_____|_|   \____)_| |_|
                            (_____|                                     
                                                            by:foyaga
        """)
        parser = argparse.ArgumentParser()
        parser.add_argument('-f', dest='file', help='导入文件批量查询')
        parser.add_argument('-t', dest='target', help='指定公司名称查询')
        args = parser.parse_args()
        return args, parser


if __name__ == '__main__':
    Scan = EnterInfoSearch()
    args, parser = Scan.banner()
    nowTime = datetime.now().strftime("%Y%m%d%H%M%S")
    if args.file:
        Scan.is_rp = False
        file = args.file
        with open(file, "r", encoding='UTF-8') as files:
            file_data = files.readlines()  # 读取文件
            for company in file_data:
                company = company.strip('\n')
                Scan.run(nowTime, company)
                sleep(5)

    elif args.target:
        Scan.run(nowTime, args.target)
    else:
        parser.parse_args(["-h"])
